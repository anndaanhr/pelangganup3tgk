"""
Script import Excel untuk SEMUA data (hampir 1 juta per tahun)
Menggunakan batch processing untuk efisiensi memory dan performa
"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import pandas as pd
from sqlalchemy import text
from database import engine
from models import Base
from dotenv import load_dotenv
import time

load_dotenv()

# Mapping kolom bulanan
MONTH_MAP_2024 = {
    '2023-12-01 00:00:00': 'dec_2023',
    '2024-01-01 00:00:00': 'jan_2024',
    '2024-02-01 00:00:00': 'feb_2024',
    '2024-03-01 00:00:00': 'mar_2024',
    '2024-04-01 00:00:00': 'apr_2024',
    '2024-05-01 00:00:00': 'may_2024',
    '2024-06-01 00:00:00': 'jun_2024',
    '2024-07-01 00:00:00': 'jul_2024',
    '2024-08-01 00:00:00': 'aug_2024',
    '2024-09-01 00:00:00': 'sep_2024',
    '2024-10-01 00:00:00': 'oct_2024',
    '2024-11-01 00:00:00': 'nov_2024',
    '2024-12-01 00:00:00': 'dec_2024',
}

MONTH_MAP_2025 = {
    '2024-12-01 00:00:00': 'dec_2024',
    '2025-01-01 00:00:00': 'jan_2025',
    '2025-02-01 00:00:00': 'feb_2025',
    '2025-03-01 00:00:00': 'mar_2025',
    '2025-04-01 00:00:00': 'apr_2025',
    '2025-05-01 00:00:00': 'may_2025',
    '2025-06-01 00:00:00': 'jun_2025',
    '2025-07-01 00:00:00': 'jul_2025',
    '2025-08-01 00:00:00': 'aug_2025',
    '2025-09-01 00:00:00': 'sep_2025',
    '2025-10-01 00:00:00': 'oct_2025',
    '2025-11-01 00:00:00': 'nov_2025',
    '2025-12-01 00:00:00': 'dec_2025',
}

BATCH_SIZE = 5000  # Process 5000 rows at a time

def escape_sql(s):
    """Escape string untuk SQL"""
    if pd.isna(s) or s is None:
        return ''
    return str(s).replace("'", "''")

def get_month_values(row, year):
    """Ambil nilai bulanan dari row"""
    month_map = MONTH_MAP_2024 if year == 2024 else MONTH_MAP_2025
    month_values = {}
    
    for col in row.index:
        col_str = str(col)
        if col_str in month_map:
            val = row[col]
            if pd.notna(val):
                try:
                    month_values[month_map[col_str]] = float(val)
                except (ValueError, TypeError):
                    month_values[month_map[col_str]] = None
            else:
                month_values[month_map[col_str]] = None
        elif isinstance(col, pd.Timestamp):
            col_str_ts = str(col)
            if col_str_ts in month_map:
                val = row[col]
                if pd.notna(val):
                    try:
                        month_values[month_map[col_str_ts]] = float(val)
                    except (ValueError, TypeError):
                        month_values[month_map[col_str_ts]] = None
                else:
                    month_values[month_map[col_str_ts]] = None
    
    return month_values

def build_insert_sql(row, year):
    """Build SQL INSERT statement untuk satu row"""
    try:
        idpel = int(row['IDPEL']) if pd.notna(row['IDPEL']) else None
        if not idpel:
            return None
        
        month_vals = get_month_values(row, year)
        
        # Build kolom dan values
        base_cols = ['idpel', 'unitup', 'nama', 'alamat', 'tarif', 'daya', 'kddk', 'cater', 'penyulang', 'gardu', 'jenis', 'layanan', 'kd_proses']
        base_vals = [
            str(idpel),
            str(int(row['UNITUP'])) if pd.notna(row['UNITUP']) else 'NULL',
            f"'{escape_sql(row['NAMA'])}'" if pd.notna(row['NAMA']) else 'NULL',
            f"'{escape_sql(row['ALAMAT'])}'" if pd.notna(row['ALAMAT']) else 'NULL',
            f"'{escape_sql(row['TARIF'])}'" if pd.notna(row['TARIF']) else 'NULL',
            str(int(row['DAYA'])) if pd.notna(row['DAYA']) else 'NULL',
            f"'{escape_sql(row['KDDK'])}'" if pd.notna(row['KDDK']) else 'NULL',
            f"'{escape_sql(row['CATER'])}'" if 'CATER' in row.index and pd.notna(row['CATER']) else 'NULL',
            f"'{escape_sql(row['PENYULANG'])}'" if 'PENYULANG' in row.index and pd.notna(row['PENYULANG']) else 'NULL',
            f"'{escape_sql(row['GARDU'])}'" if pd.notna(row['GARDU']) else 'NULL',
            f"'{escape_sql(row['JENIS'])}'" if pd.notna(row['JENIS']) else 'NULL',
            f"'{escape_sql(row['LAYANAN'])}'" if pd.notna(row['LAYANAN']) else 'NULL',
            f"'{escape_sql(row['KD PROSES'])}'" if pd.notna(row['KD PROSES']) else 'NULL'
        ]
        
        # Tambahkan kolom bulanan
        if year == 2024:
            month_cols = ['dec_2023', 'jan_2024', 'feb_2024', 'mar_2024', 'apr_2024', 'may_2024',
                          'jun_2024', 'jul_2024', 'aug_2024', 'sep_2024', 'oct_2024', 'nov_2024', 'dec_2024']
        else:
            month_cols = ['dec_2024', 'jan_2025', 'feb_2025', 'mar_2025', 'apr_2025', 'may_2025',
                          'jun_2025', 'jul_2025', 'aug_2025', 'sep_2025', 'oct_2025', 'nov_2025', 'dec_2025']
        
        month_vals_list = []
        for col in month_cols:
            if col in month_vals and month_vals[col] is not None:
                month_vals_list.append(str(month_vals[col]))
            else:
                month_vals_list.append('NULL')
        
        all_cols = ', '.join(base_cols + month_cols)
        all_vals = ', '.join(base_vals + month_vals_list)
        
        # Tambahkan MERK KWH untuk 2024
        if year == 2024 and 'MERK KWH' in row.index and pd.notna(row['MERK KWH']):
            all_cols += ', merk_kwh'
            all_vals += f", '{escape_sql(row['MERK KWH'])}'"
        
        # Build UPDATE clause
        update_cols = base_cols[1:] + month_cols
        if year == 2024 and 'MERK KWH' in row.index and pd.notna(row['MERK KWH']):
            update_cols.append('merk_kwh')
        
        update_clause = ", ".join([f"{col} = EXCLUDED.{col}" for col in update_cols])
        
        table_name = 'customers_2024' if year == 2024 else 'customers_2025'
        sql = f"INSERT INTO {table_name} ({all_cols}) VALUES ({all_vals}) ON CONFLICT (idpel) DO UPDATE SET {update_clause};"
        
        return sql, idpel
    except Exception as e:
        return None, None

def import_full():
    """Import semua data dengan batch processing"""
    print("=" * 70)
    print("IMPORT DATA FULL - SEMUA DATA (Hampir 1 Juta per Tahun)")
    print("=" * 70)
    
    project_root = Path(__file__).resolve().parent.parent.parent
    excel_2024 = project_root / "data" / "JUAL PERPELANGGAN 2024 BL.xlsx"
    excel_2025 = project_root / "data" / "JUAL PERPELANGGAN 2025 BL.xlsx"
    
    # Create tables
    print("\n[1/6] Creating/updating database tables...")
    Base.metadata.create_all(bind=engine)
    
    # Get total rows untuk progress tracking
    print("\n[2/6] Counting total rows in Excel files...")
    try:
        import openpyxl
        wb_2024 = openpyxl.load_workbook(excel_2024, read_only=True)
        ws_2024 = wb_2024.active
        total_rows_2024 = ws_2024.max_row - 1  # Exclude header
        wb_2024.close()
        
        wb_2025 = openpyxl.load_workbook(excel_2025, read_only=True)
        ws_2025 = wb_2025.active
        total_rows_2025 = ws_2025.max_row - 1  # Exclude header
        wb_2025.close()
        
        print(f"  Total rows in 2024 file: {total_rows_2024:,}")
        print(f"  Total rows in 2025 file: {total_rows_2025:,}")
    except Exception as e:
        print(f"  Warning: Could not count rows exactly: {e}")
        print("  Will process all available data...")
        total_rows_2024 = None
        total_rows_2025 = None
    
    # Import 2024 dengan batch processing
    print("\n[3/6] Importing 2024 data (batch processing)...")
    conn = engine.connect()
    trans = conn.begin()
    
    try:
        idpel_list_2024 = set()
        batch_count = 0
        total_inserted = 0
        start_time = time.time()
        skip_rows = 1  # Skip header
        
        # Baca Excel dalam batch menggunakan skiprows dan nrows
        while True:
            try:
                chunk_df = pd.read_excel(
                    excel_2024, 
                    skiprows=range(1, skip_rows + 1),  # Skip header + rows yang sudah diproses
                    nrows=BATCH_SIZE,
                    engine='openpyxl'
                )
                
                if chunk_df.empty:
                    break
                
                batch_count += 1
                batch_inserted = 0
                
                for idx, row in chunk_df.iterrows():
                    sql, idpel = build_insert_sql(row, 2024)
                    if sql and idpel:
                        try:
                            conn.execute(text(sql))
                            idpel_list_2024.add(idpel)
                            batch_inserted += 1
                            total_inserted += 1
                        except Exception as e:
                            # Skip duplicate atau error lain
                            if 'duplicate key' not in str(e).lower():
                                pass  # Silent skip untuk performa
                
                skip_rows += len(chunk_df)
                
                # Commit setiap 10 batch untuk menghindari transaction terlalu besar
                if batch_count % 10 == 0:  # Commit setiap 10 batch (50k rows)
                    trans.commit()
                    trans = conn.begin()
                
                # Progress update
                elapsed = time.time() - start_time
                if total_rows_2024:
                    progress = (total_inserted / total_rows_2024) * 100
                    print(f"  Progress: {total_inserted:,} / {total_rows_2024:,} ({progress:.1f}%) | "
                          f"Batch: {batch_count} | Time: {elapsed:.1f}s", end='\r')
                else:
                    print(f"  Processed: {total_inserted:,} rows | Batch: {batch_count} | Time: {elapsed:.1f}s", end='\r')
                
                # Jika batch lebih kecil dari BATCH_SIZE, berarti sudah sampai akhir
                if len(chunk_df) < BATCH_SIZE:
                    break
                    
            except Exception as e:
                if 'No columns to parse' in str(e) or 'EmptyDataError' in str(e):
                    break
                raise
        
        # Final commit
        trans.commit()
        elapsed = time.time() - start_time
        print(f"\n  [SUCCESS] Inserted {len(idpel_list_2024):,} records for 2024 in {elapsed:.1f} seconds")
        
    except Exception as e:
        trans.rollback()
        print(f"\n  [ERROR] Error importing 2024: {e}")
        import traceback
        traceback.print_exc()
        raise
    finally:
        conn.close()
    
    # Import 2025 dengan batch processing
    print("\n[4/6] Importing 2025 data (batch processing)...")
    conn = engine.connect()
    trans = conn.begin()
    
    try:
        idpel_list_2025 = set()
        batch_count = 0
        total_inserted = 0
        start_time = time.time()
        skip_rows = 1  # Skip header
        
        # Baca Excel dalam batch menggunakan skiprows dan nrows
        while True:
            try:
                chunk_df = pd.read_excel(
                    excel_2025, 
                    skiprows=range(1, skip_rows + 1),  # Skip header + rows yang sudah diproses
                    nrows=BATCH_SIZE,
                    engine='openpyxl'
                )
                
                if chunk_df.empty:
                    break
                
                batch_count += 1
                batch_inserted = 0
                
                for idx, row in chunk_df.iterrows():
                    sql, idpel = build_insert_sql(row, 2025)
                    if sql and idpel:
                        try:
                            conn.execute(text(sql))
                            idpel_list_2025.add(idpel)
                            batch_inserted += 1
                            total_inserted += 1
                        except Exception as e:
                            # Skip duplicate atau error lain
                            if 'duplicate key' not in str(e).lower():
                                pass  # Silent skip untuk performa
                
                skip_rows += len(chunk_df)
                
                # Commit setiap 10 batch untuk menghindari transaction terlalu besar
                if batch_count % 10 == 0:  # Commit setiap 10 batch (50k rows)
                    trans.commit()
                    trans = conn.begin()
                
                # Progress update
                elapsed = time.time() - start_time
                if total_rows_2025:
                    progress = (total_inserted / total_rows_2025) * 100
                    print(f"  Progress: {total_inserted:,} / {total_rows_2025:,} ({progress:.1f}%) | "
                          f"Batch: {batch_count} | Time: {elapsed:.1f}s", end='\r')
                else:
                    print(f"  Processed: {total_inserted:,} rows | Batch: {batch_count} | Time: {elapsed:.1f}s", end='\r')
                
                # Jika batch lebih kecil dari BATCH_SIZE, berarti sudah sampai akhir
                if len(chunk_df) < BATCH_SIZE:
                    break
                    
            except Exception as e:
                if 'No columns to parse' in str(e) or 'EmptyDataError' in str(e):
                    break
                raise
        
        # Final commit
        trans.commit()
        elapsed = time.time() - start_time
        print(f"\n  [SUCCESS] Inserted {len(idpel_list_2025):,} records for 2025 in {elapsed:.1f} seconds")
        
    except Exception as e:
        trans.rollback()
        print(f"\n  [ERROR] Error importing 2025: {e}")
        import traceback
        traceback.print_exc()
        raise
    finally:
        conn.close()
    
    # Verifikasi konsistensi
    print("\n[5/6] Verifying data consistency...")
    common_imported = idpel_list_2024 & idpel_list_2025
    new_imported = idpel_list_2025 - idpel_list_2024
    lost_imported = idpel_list_2024 - idpel_list_2025
    
    print("\n" + "=" * 70)
    print("VERIFICATION RESULTS")
    print("=" * 70)
    print(f"Total records 2024: {len(idpel_list_2024):,}")
    print(f"Total records 2025: {len(idpel_list_2025):,}")
    print(f"\nExisting customers (in both years): {len(common_imported):,}")
    print(f"New customers (only in 2025): {len(new_imported):,}")
    print(f"Lost customers (only in 2024): {len(lost_imported):,}")
    
    # Final verification dari database
    print("\n[6/6] Final database verification...")
    conn = engine.connect()
    try:
        from sqlalchemy import func
        from models import Customer2024, Customer2025
        
        db_count_2024 = conn.execute(text("SELECT COUNT(*) FROM customers_2024")).scalar()
        db_count_2025 = conn.execute(text("SELECT COUNT(*) FROM customers_2025")).scalar()
        
        print(f"Database records 2024: {db_count_2024:,}")
        print(f"Database records 2025: {db_count_2025:,}")
    finally:
        conn.close()
    
    print("\n" + "=" * 70)
    print("[SUCCESS] FULL IMPORT COMPLETED SUCCESSFULLY!")
    print("=" * 70)
    print("Ready to use! Open http://localhost:3000")
    print("=" * 70)

if __name__ == "__main__":
    import_full()

